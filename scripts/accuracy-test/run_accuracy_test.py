"""
Calls Clear AI POST /classifications for every unique value in column A
("Description") of the anonymized commercial-invoice xlsx, writes the chosen
HS code into col C and the candidate codes (chosen + alternatives) into col D,
flags matches in cols E/F, and emits an accuracy report.

Run from repo root:
    python3 scripts/accuracy-test/run_accuracy_test.py
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from threading import Lock

import openpyxl
import requests

DEFAULT_INPUT = (
    "/Users/asma/Desktop/Customs AI/sharepoint/sample_data/"
    "sample_input_commercial_invoice/client_commercial_invoices_sample2_anonymized.xlsx"
)
DEFAULT_API = "http://localhost:3000/classifications"
OUT_DIR = Path(__file__).parent / "out"


def call_describe(session: requests.Session, api: str, description: str, timeout: int) -> dict:
    try:
        r = session.post(api, json={"description": description}, timeout=timeout)
        if r.status_code != 200:
            return {"error": f"http_{r.status_code}", "body": r.text[:200]}
        return r.json()
    except requests.RequestException as e:
        return {"error": "request_exception", "detail": str(e)}


def extract_codes(resp: dict) -> tuple[str, str]:
    """Return (selected_code, candidate_codes_csv).

    candidate_codes = chosen first, then `alternatives[].code`.
    """
    if "error" in resp:
        return "", f"ERROR:{resp['error']}"
    chosen = (resp.get("result") or {}).get("code") or ""
    alts = [a.get("code") for a in (resp.get("alternatives") or []) if a.get("code")]
    candidates = [c for c in [chosen, *alts] if c]
    # de-dupe preserving order
    seen: set[str] = set()
    uniq = [c for c in candidates if not (c in seen or seen.add(c))]
    return chosen, ",".join(uniq)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default=DEFAULT_INPUT)
    ap.add_argument("--api", default=DEFAULT_API)
    ap.add_argument("--concurrency", type=int, default=8)
    ap.add_argument("--timeout", type=int, default=120)
    ap.add_argument("--limit-unique", type=int, default=0,
                    help="If >0, only classify the first N unique descriptions (smoke test).")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    src = Path(args.input)
    if not src.exists():
        print(f"input not found: {src}", file=sys.stderr)
        sys.exit(1)

    print(f"[1/4] loading workbook: {src.name}", flush=True)
    t0 = time.time()
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    print(f"      sheet={ws.title} rows={ws.max_row} cols={ws.max_column} loaded in {time.time()-t0:.1f}s",
          flush=True)

    # Header sanity
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"      headers={headers}", flush=True)
    assert headers[0] == "Description", "col A must be 'Description'"

    # Collect unique descriptions
    print("[2/4] collecting unique descriptions...", flush=True)
    unique: dict[str, list[int]] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        key = str(v).strip()
        if not key:
            continue
        unique.setdefault(key, []).append(r)
    print(f"      {len(unique)} unique descriptions across {sum(len(v) for v in unique.values())} rows",
          flush=True)

    keys = list(unique.keys())
    if args.limit_unique > 0:
        keys = keys[: args.limit_unique]
        print(f"      LIMIT applied: only classifying first {len(keys)} unique descriptions", flush=True)

    # Classify
    print(f"[3/4] classifying via {args.api} (concurrency={args.concurrency})...", flush=True)
    results: dict[str, tuple[str, str]] = {}
    raw_log_path = OUT_DIR / "raw_responses.jsonl"
    raw_log = raw_log_path.open("w", encoding="utf-8")
    log_lock = Lock()

    progress_every = max(50, len(keys) // 200)
    done = 0
    errors = 0
    t_start = time.time()

    session = requests.Session()
    with ThreadPoolExecutor(max_workers=args.concurrency) as ex:
        futures = {ex.submit(call_describe, session, args.api, k, args.timeout): k for k in keys}
        for fut in as_completed(futures):
            k = futures[fut]
            resp = fut.result()
            sel, cand = extract_codes(resp)
            results[k] = (sel, cand)
            with log_lock:
                raw_log.write(json.dumps({"description": k, "response": resp}, ensure_ascii=False) + "\n")
            if "error" in resp or not sel:
                errors += 1
            done += 1
            if done % progress_every == 0 or done == len(keys):
                elapsed = time.time() - t_start
                rate = done / elapsed if elapsed > 0 else 0
                remaining = (len(keys) - done) / rate if rate > 0 else 0
                print(f"      [{done}/{len(keys)}] {rate:.2f} req/s "
                      f"errors={errors} eta={remaining/60:.1f}m", flush=True)

    raw_log.close()
    print(f"      classification done in {(time.time()-t_start)/60:.1f}m, errors={errors}", flush=True)

    # Write back to xlsx + accuracy
    print("[4/4] writing results & computing accuracy...", flush=True)
    total = 0
    rows_with_pred = 0
    rows_with_truth = 0
    full_match = 0
    hs10_match = 0
    hs6_match = 0
    in_candidates_match = 0

    for desc, row_indices in unique.items():
        sel, cand = results.get(desc, ("", ""))
        cand_set = set(c for c in cand.split(",") if c and not c.startswith("ERROR:"))
        for r in row_indices:
            total += 1
            truth_raw = ws.cell(row=r, column=2).value
            truth = str(truth_raw).strip() if truth_raw is not None else ""
            # Excel may show 12-digit codes as numbers; force string of digits
            if truth and truth.endswith(".0"):
                truth = truth[:-2]

            ws.cell(row=r, column=3).value = sel or None
            ws.cell(row=r, column=4).value = cand or None

            sel_match = bool(sel) and bool(truth) and sel == truth
            cand_match = bool(truth) and truth in cand_set
            ws.cell(row=r, column=6).value = bool(sel_match)
            ws.cell(row=r, column=7).value = bool(cand_match)

            if sel:
                rows_with_pred += 1
            if truth:
                rows_with_truth += 1
            if sel and truth:
                if sel == truth:
                    full_match += 1
                if sel[:10] == truth[:10]:
                    hs10_match += 1
                if sel[:6] == truth[:6]:
                    hs6_match += 1
                if cand_match:
                    in_candidates_match += 1

    out_xlsx = OUT_DIR / "classified.xlsx"
    wb.save(out_xlsx)
    print(f"      saved: {out_xlsx}", flush=True)

    # Report
    def pct(n: int, d: int) -> str:
        return f"{(100.0*n/d):.2f}%" if d else "n/a"

    denom_predtruth = min(rows_with_pred, rows_with_truth)
    # Strict denominator: rows where we have BOTH a prediction and a ground truth.
    strict_denom = sum(
        1 for desc, ridx in unique.items()
        for r in ridx
        if results.get(desc, ("", ""))[0]
        and (str(ws.cell(row=r, column=2).value or "").strip() not in ("", "None"))
    )

    report_lines = [
        "Clear AI accuracy report",
        f"input: {src}",
        f"api:   {args.api}",
        f"unique descriptions classified: {len(keys)}  (errors: {errors})",
        f"total rows: {total}",
        f"rows with prediction: {rows_with_pred}",
        f"rows with ground truth: {rows_with_truth}",
        f"rows with both:        {strict_denom}",
        "",
        "Accuracy (denominator = rows with both prediction and ground truth):",
        f"  exact 12-digit match:        {full_match} / {strict_denom}  ({pct(full_match, strict_denom)})",
        f"  10-digit prefix match:       {hs10_match} / {strict_denom}  ({pct(hs10_match, strict_denom)})",
        f"  HS-6 prefix match:           {hs6_match} / {strict_denom}  ({pct(hs6_match, strict_denom)})",
        f"  truth in candidate set:      {in_candidates_match} / {strict_denom}  ({pct(in_candidates_match, strict_denom)})",
    ]
    report = "\n".join(report_lines)
    print("\n" + report)

    (OUT_DIR / "accuracy_report.txt").write_text(report + "\n", encoding="utf-8")
    print(f"\n      report: {OUT_DIR/'accuracy_report.txt'}")
    print(f"      raw responses: {raw_log_path}")


if __name__ == "__main__":
    main()
