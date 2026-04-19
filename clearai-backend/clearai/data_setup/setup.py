"""
One-time database setup.

Creates the SQLite schema and loads every mapping xlsx from `data/` into the
appropriate table. Idempotent — safe to re-run. Drops & recreates tables each
time so schema migrations stay simple during V1.

Run:
    python -m clearai.data_setup.setup

Verify:
    sqlite3 clear_ai.db "SELECT name FROM sqlite_master WHERE type='table';"
"""

from __future__ import annotations

import json
import logging
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

import openpyxl

from clearai import config

logger = logging.getLogger("clearai.data_setup.setup")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-7s %(message)s")


# ---------------------------------------------------------------------------
# Source file paths
# ---------------------------------------------------------------------------
DATA_DIR = config.DATA_DIR
FILE_LEDGER = DATA_DIR / "Naqel_HS_code_mapping_lookup.xlsx"
FILE_MASTER = DATA_DIR / "Zatca Tariff codes.xlsx"
FILE_FIELDS = DATA_DIR / "Naqel (Fields details + Mapping data).xlsx"


# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------
SCHEMA = """
-- 1. HS decision ledger — seeded from Naqel's historical human-verified mappings.
--    `client_id` is nullable: the seed data is global (no client column).
--    Runtime ledger entries (from review write-back) will set client_id.
DROP TABLE IF EXISTS hs_decision_ledger;
CREATE TABLE hs_decision_ledger (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    client_id       TEXT,
    raw_code        TEXT NOT NULL,
    verified_code   TEXT NOT NULL,
    arabic_name     TEXT,
    unit_per_price  INTEGER,
    provenance      TEXT DEFAULT 'human_verified',
    created_at      TEXT
);
CREATE UNIQUE INDEX idx_ledger_clientraw ON hs_decision_ledger(
    COALESCE(client_id, ''), raw_code
);
CREATE INDEX idx_ledger_raw ON hs_decision_ledger(raw_code);

-- 2. HS code master — full ZATCA tariff (~19k rows).
DROP TABLE IF EXISTS hs_code_master;
CREATE TABLE hs_code_master (
    hs_code          TEXT PRIMARY KEY,         -- 12-digit zero-padded
    arabic_name      TEXT,
    description_en   TEXT,                     -- source for FAISS embedding
    duty_rate_text   TEXT,                     -- raw duty rate string ("5%", "معفاة", "")
    duty_rate_pct    REAL,                     -- parsed percentage, NULL if not numeric
    procedures       TEXT,
    effective_date   TEXT
);
CREATE INDEX idx_master_prefix6 ON hs_code_master(SUBSTR(hs_code, 1, 6));
CREATE INDEX idx_master_prefix4 ON hs_code_master(SUBSTR(hs_code, 1, 4));

-- 3. Currency mapping — InfoTrack → Tabdul.
DROP TABLE IF EXISTS currency_mapping;
CREATE TABLE currency_mapping (
    infotrack_currency_id   INTEGER PRIMARY KEY,
    tabdul_currency_id      INTEGER NOT NULL,
    iso_code                TEXT                -- SAR, AED, USD, …
);
CREATE INDEX idx_currency_iso ON currency_mapping(iso_code);

-- 4. City mapping bridge — 2-step lookup.
--    DestinationStationID → info_city_id → tabdul_city_id → Tabdul City details.
DROP TABLE IF EXISTS city_mapping_bridge;
CREATE TABLE city_mapping_bridge (
    tabdul_city_id   INTEGER,
    info_city_id     INTEGER
);
CREATE INDEX idx_city_info ON city_mapping_bridge(info_city_id);
CREATE INDEX idx_city_tabdul ON city_mapping_bridge(tabdul_city_id);

-- 5. Tabdul City — city master from Tabadul system.
--    NOTE: CITY_CD is NOT globally unique; it's a within-country serial (cycles
--    from 1 per country). The real unique key is (city_cd, ctry_cd).
DROP TABLE IF EXISTS tabdul_city;
CREATE TABLE tabdul_city (
    city_cd          INTEGER NOT NULL,
    ctry_cd          INTEGER NOT NULL,
    city_arb_name    TEXT,
    city_eng_name    TEXT,
    city_intl_cd     TEXT,
    mins_comm_cd     TEXT,
    city_status      INTEGER,
    PRIMARY KEY (city_cd, ctry_cd)
);
CREATE INDEX idx_tabdul_city_intl ON tabdul_city(city_intl_cd);
CREATE INDEX idx_tabdul_city_eng ON tabdul_city(city_eng_name);

-- 6. Source company mapping — (client_id, cust_reg_port_code) → source company.
--    Fallback used when no row matches: (client_id=-1, cust_reg_port_code=23) → 'ناقل'.
DROP TABLE IF EXISTS source_company_mapping;
CREATE TABLE source_company_mapping (
    id                       INTEGER PRIMARY KEY AUTOINCREMENT,
    source_company_name      TEXT NOT NULL,
    source_company_no        TEXT,                    -- kept as TEXT: observed non-numeric values ("QA Test")
    client_id                INTEGER,
    cust_reg_port_code       INTEGER
);
CREATE UNIQUE INDEX idx_source_clientport ON source_company_mapping(client_id, cust_reg_port_code);

-- 7. Country of origin by client.
DROP TABLE IF EXISTS country_origin_mapping;
CREATE TABLE country_origin_mapping (
    client_id         INTEGER PRIMARY KEY,
    country_origin    INTEGER NOT NULL        -- Tabadul CountryCode numeric
);

-- 8. Country code reference — translate numeric → ISO INTLCODE.
--    Required to emit <countryOfOrigin>CN</countryOfOrigin> etc.
DROP TABLE IF EXISTS country_code;
CREATE TABLE country_code (
    country_code     INTEGER PRIMARY KEY,
    name_arabic      TEXT,
    name_english     TEXT,
    intl_code        TEXT                      -- 2-letter ISO
);
CREATE INDEX idx_country_intl ON country_code(intl_code);

-- Meta table — track setup runs.
DROP TABLE IF EXISTS setup_meta;
CREATE TABLE setup_meta (
    key         TEXT PRIMARY KEY,
    value       TEXT,
    updated_at  TEXT
);
"""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _iter_rows(path: Path, sheet: str) -> Iterator[tuple]:
    """Stream rows from a named sheet, skipping the header row."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise RuntimeError(f"{path.name}: sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    row_iter = ws.iter_rows(values_only=True)
    next(row_iter, None)  # skip header
    try:
        for row in row_iter:
            yield row
    finally:
        wb.close()


def _normalize_code(raw) -> str:
    """Strip non-digits and zero-pad to 12 digits. Empty string on None/invalid."""
    if raw is None:
        return ""
    digits = re.sub(r"[^0-9]", "", str(raw))
    if not digits:
        return ""
    return digits.zfill(12) if len(digits) <= 12 else digits[:12]


_DUTY_PCT_RE = re.compile(r"(\d+(?:\.\d+)?)\s*%")


def _parse_duty_rate(text) -> float | None:
    """Extract numeric percentage from Arabic/English duty text. None if non-numeric."""
    if text is None:
        return None
    s = str(text).strip()
    if not s or s == "معفاة" or s.lower() == "exempted":
        return 0.0
    m = _DUTY_PCT_RE.search(s)
    if m:
        return float(m.group(1))
    try:
        return float(s)
    except ValueError:
        return None


def _clean(v):
    """Strip whitespace on strings, keep everything else."""
    if isinstance(v, str):
        return v.strip()
    return v


# ---------------------------------------------------------------------------
# Loaders (one per source file/sheet)
# ---------------------------------------------------------------------------
def load_ledger(conn: sqlite3.Connection) -> int:
    """Naqel_HS_code_mapping_lookup.xlsx → hs_decision_ledger."""
    now = datetime.now(timezone.utc).isoformat()
    rows = []
    for r in _iter_rows(FILE_LEDGER, "Sheet1"):
        raw, verified, unit_per_price, arabic = r[0], r[1], r[2], r[3]
        raw_norm = _normalize_code(raw)
        ver_norm = _normalize_code(verified)
        if not raw_norm or not ver_norm:
            continue
        rows.append(
            (
                None,  # client_id — global seed entries
                raw_norm,
                ver_norm,
                _clean(arabic),
                int(unit_per_price) if unit_per_price is not None else None,
                "naqel_seed",
                now,
            )
        )
    conn.executemany(
        """INSERT OR IGNORE INTO hs_decision_ledger
           (client_id, raw_code, verified_code, arabic_name, unit_per_price, provenance, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        rows,
    )
    return len(rows)


def load_hs_master(conn: sqlite3.Connection) -> int:
    """Zatca Tariff codes.xlsx (sheet: Grid) → hs_code_master."""
    rows = []
    seen: set[str] = set()
    for r in _iter_rows(FILE_MASTER, "Grid"):
        hs_code = _normalize_code(r[0])
        if not hs_code or hs_code in seen:
            continue
        seen.add(hs_code)
        rows.append(
            (
                hs_code,
                _clean(r[1]),                    # arabic_name
                _clean(r[2]),                    # description_en
                _clean(r[4]) or None,            # duty_rate_text (English)
                _parse_duty_rate(r[4]),          # duty_rate_pct
                _clean(r[5]) or None,            # procedures
                _clean(r[6]) or None,            # effective_date
            )
        )
    conn.executemany(
        """INSERT OR REPLACE INTO hs_code_master
           (hs_code, arabic_name, description_en, duty_rate_text, duty_rate_pct, procedures, effective_date)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        rows,
    )
    return len(rows)


def load_currency(conn: sqlite3.Connection) -> int:
    """Naqel (Fields details).xlsx sheet CurrencyMapping → currency_mapping."""
    rows = []
    for r in _iter_rows(FILE_FIELDS, "CurrencyMapping"):
        info_id, tabdul_id, iso = r[0], r[1], r[2]
        if info_id is None or tabdul_id is None:
            continue
        rows.append((int(info_id), int(tabdul_id), _clean(iso)))
    conn.executemany(
        "INSERT OR REPLACE INTO currency_mapping VALUES (?, ?, ?)",
        rows,
    )
    return len(rows)


def load_city_bridge(conn: sqlite3.Connection) -> int:
    """Naqel (Fields details).xlsx sheet CityMaping → city_mapping_bridge."""
    rows = []
    for r in _iter_rows(FILE_FIELDS, "CityMaping"):
        tabdul_id, info_id = r[0], r[1]
        if tabdul_id is None and info_id is None:
            continue
        rows.append(
            (
                int(tabdul_id) if tabdul_id is not None else None,
                int(info_id) if info_id is not None else None,
            )
        )
    conn.executemany(
        "INSERT INTO city_mapping_bridge VALUES (?, ?)",
        rows,
    )
    return len(rows)


def load_tabdul_city(conn: sqlite3.Connection) -> int:
    """Naqel (Fields details).xlsx sheet 'Tabdul City' → tabdul_city.

    Note: primary key is the composite (city_cd, ctry_cd) because CITY_CD
    cycles 1..N within each country.
    """
    rows = []
    seen: set[tuple[int, int]] = set()
    for r in _iter_rows(FILE_FIELDS, "Tabdul City"):
        city_cd, ctry_cd = r[0], r[6]
        if city_cd is None or ctry_cd is None:
            continue
        key = (int(city_cd), int(ctry_cd))
        if key in seen:
            continue
        seen.add(key)
        mins = r[4]
        if isinstance(mins, str) and mins.upper() == "NULL":
            mins = None
        rows.append(
            (
                int(city_cd),
                int(ctry_cd),
                _clean(r[1]),
                _clean(r[2]),
                _clean(r[3]),
                mins,
                int(r[5]) if r[5] is not None else None,
            )
        )
    conn.executemany(
        """INSERT OR REPLACE INTO tabdul_city
           (city_cd, ctry_cd, city_arb_name, city_eng_name, city_intl_cd, mins_comm_cd, city_status)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        rows,
    )
    return len(rows)


def load_source_company(conn: sqlite3.Connection) -> int:
    """Naqel (Fields details).xlsx sheet SourceCompanyPortMaping → source_company_mapping."""
    rows = []
    for r in _iter_rows(FILE_FIELDS, "SourceCompanyPortMaping"):
        name, no, client, port = r[0], r[1], r[2], r[3]
        if not name:
            continue
        # source_company_no is kept as TEXT because one row in the
        # received data is the string "QA Test".
        rows.append(
            (
                _clean(name),
                str(no).strip() if no is not None else None,
                int(client) if client is not None else None,
                int(port) if port is not None else None,
            )
        )
    conn.executemany(
        """INSERT INTO source_company_mapping
           (source_company_name, source_company_no, client_id, cust_reg_port_code)
           VALUES (?, ?, ?, ?)
           ON CONFLICT(client_id, cust_reg_port_code) DO UPDATE SET
             source_company_name = excluded.source_company_name,
             source_company_no = excluded.source_company_no""",
        rows,
    )
    return len(rows)


def load_country_origin(conn: sqlite3.Connection) -> int:
    """Naqel (Fields details).xlsx sheet CountryOfOriginClientMapping → country_origin_mapping."""
    rows = []
    for r in _iter_rows(FILE_FIELDS, "CountryOfOriginClientMapping"):
        client_id, origin = r[0], r[1]
        if client_id is None or origin is None:
            continue
        rows.append((int(client_id), int(origin)))
    conn.executemany(
        "INSERT OR REPLACE INTO country_origin_mapping VALUES (?, ?)",
        rows,
    )
    return len(rows)


def load_country_codes(conn: sqlite3.Connection) -> int:
    """Naqel (Fields details).xlsx sheet Tabadul CountryCode → country_code."""
    rows = []
    for r in _iter_rows(FILE_FIELDS, "Tabadul CountryCode"):
        code, arabic, english, intl = r[0], r[1], r[2], r[3]
        if code is None:
            continue
        rows.append((int(code), _clean(arabic), _clean(english), _clean(intl)))
    conn.executemany(
        "INSERT OR REPLACE INTO country_code VALUES (?, ?, ?, ?)",
        rows,
    )
    return len(rows)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def main() -> int:
    # Validate source files exist before touching the DB
    required = [FILE_LEDGER, FILE_MASTER, FILE_FIELDS]
    missing = [p for p in required if not p.is_file()]
    if missing:
        logger.error("Missing required data files:")
        for p in missing:
            logger.error("  %s", p)
        logger.error("Place the xlsx files in %s and re-run.", DATA_DIR)
        return 2

    logger.info("Opening database at %s", config.DB_PATH)
    conn = sqlite3.connect(config.DB_PATH)
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA foreign_keys = ON")

    try:
        logger.info("Creating schema…")
        conn.executescript(SCHEMA)

        loaders = [
            ("hs_decision_ledger", load_ledger),
            ("hs_code_master", load_hs_master),
            ("currency_mapping", load_currency),
            ("city_mapping_bridge", load_city_bridge),
            ("tabdul_city", load_tabdul_city),
            ("source_company_mapping", load_source_company),
            ("country_origin_mapping", load_country_origin),
            ("country_code", load_country_codes),
        ]

        counts: dict[str, int] = {}
        for name, fn in loaders:
            logger.info("Loading %s…", name)
            inserted = fn(conn)
            # Final count from table (may differ if INSERT OR IGNORE skipped dupes)
            final = conn.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0]
            counts[name] = final
            logger.info("  %s: %d rows (attempted %d)", name, final, inserted)

        now = datetime.now(timezone.utc).isoformat()
        conn.execute(
            "INSERT OR REPLACE INTO setup_meta VALUES (?, ?, ?)",
            ("last_setup_at", now, now),
        )
        conn.execute(
            "INSERT OR REPLACE INTO setup_meta VALUES (?, ?, ?)",
            ("row_counts", json.dumps(counts), now),
        )
        conn.commit()

        logger.info("=" * 60)
        logger.info("Setup complete.")
        logger.info("Row counts:")
        for k, v in counts.items():
            logger.info("  %-28s %d", k, v)

        # Sanity checks
        expected_min = {
            "hs_decision_ledger": 400,
            "hs_code_master": 18000,
            "currency_mapping": 10,
            "city_mapping_bridge": 300,
            "tabdul_city": 1000,   # ~1,084 unique (city_cd, ctry_cd); source has exact dupes
            "source_company_mapping": 150,
            "country_origin_mapping": 100,
            "country_code": 300,
        }
        failures = [
            (k, counts[k], exp) for k, exp in expected_min.items() if counts[k] < exp
        ]
        if failures:
            logger.warning("Some tables loaded fewer rows than expected:")
            for k, got, exp in failures:
                logger.warning("  %s: got %d, expected ≥%d", k, got, exp)
            return 1

        logger.info("All row counts meet expected minimums.")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
