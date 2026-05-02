#!/usr/bin/env python3
"""
build-eval-set.py — sample 500 stratified broker invoice rows for the eval suite.

Run once to produce eval/data/broker-invoices-v1.jsonl. Idempotent:
random.seed(42) makes the sample reproducible.

Usage:
    python scripts/build-eval-set.py \\
      --xlsx ../naqel-shared-data/client_commercial_invoices_sample2_anonymized.xlsx \\
      --out  eval/data/broker-invoices-v1.jsonl \\
      --total 500

Stratification (matches the broker workload survey from 100k-row sample):
    1-word    150 / 30%
    2-word    250 / 50%
    3-word     75 / 15%
    4+ word    25 /  5%
"""
from __future__ import annotations

import argparse
import json
import random
import sys
from pathlib import Path

try:
    import openpyxl  # type: ignore
except ImportError:
    sys.exit("openpyxl required; install via `pip install openpyxl`")


SEED = 42
BUCKETS = {
    "len_1":      150,
    "len_2":      250,
    "len_3":       75,
    "len_4plus":   25,
}
TOTAL = sum(BUCKETS.values())


def bucket_for(desc: str) -> str:
    """Return the length bucket for a description."""
    n = len(desc.split())
    if n == 1: return "len_1"
    if n == 2: return "len_2"
    if n == 3: return "len_3"
    return "len_4plus"


def is_valid_row(desc: str | None, code: str | None) -> bool:
    """Both fields present, code is exactly 12 digits, description non-empty."""
    if not desc or not code: return False
    code = str(code).strip()
    if len(code) != 12 or not code.isdigit(): return False
    if not str(desc).strip(): return False
    return True


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True, type=Path,
                   help="Path to the broker invoice xlsx")
    p.add_argument("--out", required=True, type=Path,
                   help="Output JSONL path")
    p.add_argument("--total", type=int, default=TOTAL,
                   help=f"Total rows (default {TOTAL}; reweights buckets proportionally)")
    args = p.parse_args()

    if not args.xlsx.is_file():
        sys.exit(f"xlsx not found: {args.xlsx}")

    # Reweight bucket sizes if --total differs from default.
    scale = args.total / TOTAL
    targets = {k: max(1, int(round(v * scale))) for k, v in BUCKETS.items()}
    print(f"Target sample sizes: {targets} (total {sum(targets.values())})")

    print(f"Reading {args.xlsx} …")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Stream all valid rows into per-bucket pools.
    pools: dict[str, list[tuple[str, str]]] = {k: [] for k in BUCKETS}
    n_total = 0
    n_invalid = 0
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        desc, code = r[0], r[1]
        if not is_valid_row(desc, code):
            n_invalid += 1
            continue
        desc = str(desc).strip()
        code = str(code).strip()
        bucket = bucket_for(desc)
        pools[bucket].append((desc, code))
        n_total += 1

    print(f"Pool sizes (after validation): {{ {', '.join(f'{k}: {len(v)}' for k, v in pools.items())} }}")
    print(f"Total valid rows surveyed: {n_total} (skipped {n_invalid} invalid)")

    # Stratified sample with deduplication on description (case-insensitive).
    random.seed(SEED)
    sampled: list[tuple[str, str, str]] = []  # (desc, code, bucket)
    for bucket, target in targets.items():
        pool = pools[bucket]
        if len(pool) < target:
            print(f"  WARN: bucket {bucket} only has {len(pool)} rows, requested {target} — taking all")
            chosen = pool[:]
        else:
            # Dedupe by lowercased description, keep first per dup-key in shuffle order
            random.shuffle(pool)
            seen: set[str] = set()
            chosen = []
            for d, c in pool:
                key = d.lower()
                if key in seen:
                    continue
                seen.add(key)
                chosen.append((d, c))
                if len(chosen) >= target:
                    break
            if len(chosen) < target:
                print(f"  WARN: bucket {bucket} after dedup only yielded {len(chosen)} unique, "
                      f"requested {target}")
        for d, c in chosen:
            sampled.append((d, c, bucket))

    # Stable id assignment: sort by bucket then by lowercased description for
    # deterministic output (so re-runs with the same seed produce byte-identical files).
    sampled.sort(key=lambda x: (list(BUCKETS).index(x[2]), x[0].lower()))

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8") as f:
        for i, (desc, code, bucket) in enumerate(sampled, 1):
            row = {
                "id": i,
                "description": desc,
                "broker_code": code,
                "broker_chapter": code[:2],
                "broker_heading": code[:4],
                "length_bucket": bucket,
                "quality": "default",
            }
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    print(f"\n✓ Wrote {len(sampled)} rows to {args.out}")
    print("Bucket breakdown of sampled set:")
    from collections import Counter
    counts = Counter(b for _, _, b in sampled)
    for bucket in BUCKETS:
        print(f"  {bucket:<12s}  {counts[bucket]:>4d}  ({100*counts[bucket]/len(sampled):.1f}%)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
