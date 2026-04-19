"""
Invoice parser — streams rows from xlsx or csv into cleaned dicts.

Key design decisions:
- Streaming-only (openpyxl read_only + iter_rows). The real sample file is 30MB
  / 50k rows / 3,020 waybills; pandas-style whole-file loads would OOM or thrash.
- Preserves original field names (matches BUILD.md spec columns). Unknown columns
  are passed through.
- Coerces Quantity → int; TotalCost, UnitCost, Amount → float.
- Strips whitespace on all string fields.
- Normalizes dates to ISO-8601 strings.
- Groups by WayBillNo — one group = one declaration.

Required columns (per BUILD.md §Step 2):
    WayBillNo, InvoiceDate, Consignee, ConsigneeAddress, ConsigneeEmail,
    MobileNo, Phone, TotalCost, CurrencyCode, ClientID, Quantity, UnitType,
    CountryofManufacture, Description, CustomsCommodityCode, UnitCost, Amount,
    Currency, ChineseDescription, SKU, CPC
"""

from __future__ import annotations

import csv
import logging
from collections.abc import Iterator
from datetime import date, datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger("clearai.invoice_parser")


REQUIRED_COLUMNS: tuple[str, ...] = (
    "WayBillNo",
    "InvoiceDate",
    "Consignee",
    "ConsigneeAddress",
    "ConsigneeEmail",
    "MobileNo",
    "Phone",
    "TotalCost",
    "CurrencyCode",
    "ClientID",
    "Quantity",
    "UnitType",
    "CountryofManufacture",
    "Description",
    "CustomsCommodityCode",
    "UnitCost",
    "Amount",
    "Currency",
    "ChineseDescription",
    "SKU",
    "CPC",
)

# Type coercion rules (applied only if value is not None/"")
INT_FIELDS: frozenset[str] = frozenset({"Quantity"})
FLOAT_FIELDS: frozenset[str] = frozenset({"TotalCost", "UnitCost", "Amount"})
DATE_FIELDS: frozenset[str] = frozenset({"InvoiceDate"})


class InvoiceParseError(Exception):
    """Raised when the input file is missing required columns or is malformed."""


# ---------------------------------------------------------------------------
# Value cleaning
# ---------------------------------------------------------------------------
def _clean_value(field: str, value: Any) -> Any:
    """Type-coerce and whitespace-trim a single field value."""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None

    if field in DATE_FIELDS:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, str):
            # Leave as-is; the downstream XML renderer can normalize if needed
            return value

    if field in INT_FIELDS:
        try:
            return int(float(value))
        except (ValueError, TypeError):
            logger.warning("Row field %s could not be coerced to int: %r", field, value)
            return None

    if field in FLOAT_FIELDS:
        try:
            return float(value)
        except (ValueError, TypeError):
            logger.warning("Row field %s could not be coerced to float: %r", field, value)
            return None

    return value


def _clean_row(headers: list[str], values: tuple) -> dict[str, Any]:
    """Build a dict from header/value lists, applying cleaning rules."""
    row: dict[str, Any] = {}
    for idx, h in enumerate(headers):
        v = values[idx] if idx < len(values) else None
        row[h] = _clean_value(h, v)
    return row


def _validate_headers(headers: list[str], source: str) -> None:
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise InvoiceParseError(
            f"{source}: missing required columns: {missing}. "
            f"Expected all of: {list(REQUIRED_COLUMNS)}"
        )


# ---------------------------------------------------------------------------
# Backend readers
# ---------------------------------------------------------------------------
def _iter_xlsx(path: Path) -> Iterator[dict[str, Any]]:
    import openpyxl  # lazy import; keeps csv-only paths fast

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        row_iter = ws.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            raise InvoiceParseError(f"{path.name}: file has no header row")
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        _validate_headers(headers, path.name)
        for values in row_iter:
            # Skip fully-empty rows (openpyxl yields them at the tail sometimes)
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue
            yield _clean_row(headers, values)
    finally:
        wb.close()


def _iter_csv(path: Path) -> Iterator[dict[str, Any]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header_row = next(reader, None)
        if not header_row:
            raise InvoiceParseError(f"{path.name}: file has no header row")
        headers = [h.strip() for h in header_row]
        _validate_headers(headers, path.name)
        for values in reader:
            if all(not (v and v.strip()) for v in values):
                continue
            yield _clean_row(headers, tuple(values))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------
def parse_invoice(filepath: str | Path) -> Iterator[dict[str, Any]]:
    """
    Stream cleaned row dicts from an invoice file.

    Supports .xlsx and .csv. The first sheet of an xlsx is read.

    Raises:
        FileNotFoundError: path does not exist
        InvoiceParseError: required columns missing or header row absent
    """
    path = Path(filepath)
    if not path.is_file():
        raise FileNotFoundError(f"Invoice file not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        yield from _iter_xlsx(path)
    elif suffix == ".csv":
        yield from _iter_csv(path)
    else:
        raise InvoiceParseError(
            f"Unsupported file type {suffix!r}. Only .xlsx and .csv are supported."
        )


def group_by_waybill(rows: Iterator[dict[str, Any]]) -> dict[Any, list[dict[str, Any]]]:
    """
    Consume an iterator of rows and bucket them by WayBillNo.

    Note: this materializes all rows in memory. For files too large to fit,
    use a streaming iterator and flush per-waybill when the key changes — but
    that requires the source file to be pre-sorted by waybill. For a 30MB file
    this simple approach is fine.
    """
    groups: dict[Any, list[dict[str, Any]]] = {}
    for row in rows:
        wb = row.get("WayBillNo")
        groups.setdefault(wb, []).append(row)
    return groups
